import io
from typing import List

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from app.core.config import get_db
from app.models.models import Teacher
from app.schemas.teacher import TeacherCreate, TeacherRead, TeacherUpdate, TeacherImportRowResult, TeacherImportResponse


router = APIRouter(prefix="/teachers", tags=["teachers"])


def _dump_payload(payload):
    """
    Support both Pydantic v1 (`dict`) and v2 (`model_dump`).
    """
    if hasattr(payload, "model_dump"):
        return payload.model_dump(exclude_unset=True)

    return payload.dict(exclude_unset=True)


def _normalise_teacher_name(name: str) -> str:
    """
    Normalise whitespace around a teacher name.

    Example:
        "  Prof. Sharma  " -> "Prof. Sharma"
    """
    return " ".join(name.strip().split())


def _find_duplicate_teacher(
    db: Session,
    teacher_name: str,
    exclude_teacher_id: int | None = None,
):
    """
    Find a teacher with the same name, ignoring case and
    surrounding/duplicate whitespace.
    """

    normalised_name = _normalise_teacher_name(teacher_name)

    query = db.query(Teacher).filter(
        func.lower(func.trim(Teacher.teacher_name))
        == normalised_name.lower()
    )

    if exclude_teacher_id is not None:
        query = query.filter(
            Teacher.teacher_id != exclude_teacher_id
        )

    return query.first()


@router.get("", response_model=List[TeacherRead])
def get_teachers(
    db: Session = Depends(get_db),
) -> List[TeacherRead]:

    teachers = (
        db.query(Teacher)
        .order_by(Teacher.teacher_id)
        .all()
    )

    return [
        TeacherRead(
            teacher_id=teacher.teacher_id,
            teacher_name=teacher.teacher_name,
            max_periods_per_day=teacher.max_periods_per_day,
        )
        for teacher in teachers
    ]


@router.get("/import/template")
def download_import_template():
    """
    Generate and return an Excel template for teacher imports.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"

    ws.append(["teacher_name", "max_periods_per_day"])
    ws.append(["Prof. Sharma", 6])

    # Auto-size columns
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=teacher_import_template.xlsx"
        },
    )


@router.post("/import", response_model=TeacherImportResponse)
def import_teachers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> TeacherImportResponse:
    """
    Bulk-import teachers from an Excel (.xlsx) file.

    Valid rows are inserted. Invalid and duplicate rows are
    skipped and reported — they do NOT block other rows.
    """

    # ── Validate file type ──────────────────────────────────
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are accepted.",
        )

    # ── Read workbook ───────────────────────────────────────
    try:
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read the Excel file. Please upload a valid .xlsx file.",
        )

    ws = wb.active
    if ws is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The workbook has no active sheet.",
        )

    rows_iter = ws.iter_rows(values_only=True)

    # ── Validate headers ────────────────────────────────────
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The Excel file is empty.",
        )

    headers = [
        str(h).strip().lower() if h is not None else ""
        for h in header_row
    ]

    required_cols = {"teacher_name", "max_periods_per_day"}
    found_cols = set(headers) & required_cols

    if found_cols != required_cols:
        missing = required_cols - found_cols
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required column(s): {', '.join(sorted(missing))}. "
                   f"Found: {', '.join(headers) or '(none)'}.",
        )

    name_idx = headers.index("teacher_name")
    periods_idx = headers.index("max_periods_per_day")

    # ── Read and validate data rows ─────────────────────────
    results: list[TeacherImportRowResult] = []
    valid_teachers: list[dict] = []
    seen_names: dict[str, int] = {}  # normalised_lower -> first row number

    for row_num_offset, row in enumerate(rows_iter):
        row_num = row_num_offset + 2  # 1-indexed, header is row 1

        # Extract raw values
        raw_name = row[name_idx] if name_idx < len(row) else None
        raw_periods = row[periods_idx] if periods_idx < len(row) else None

        # ── Validate teacher_name ───────────────────────────
        if raw_name is None or str(raw_name).strip() == "":
            results.append(TeacherImportRowResult(
                row=row_num,
                teacher_name=str(raw_name or ""),
                status="invalid",
                reason="Teacher name is empty.",
            ))
            continue

        teacher_name = _normalise_teacher_name(str(raw_name))

        if not teacher_name:
            results.append(TeacherImportRowResult(
                row=row_num,
                teacher_name=str(raw_name),
                status="invalid",
                reason="Teacher name is empty after trimming.",
            ))
            continue

        if len(teacher_name) > 100:
            results.append(TeacherImportRowResult(
                row=row_num,
                teacher_name=teacher_name[:50] + "...",
                status="invalid",
                reason="Teacher name exceeds 100 characters.",
            ))
            continue

        # ── Validate max_periods_per_day ────────────────────
        if raw_periods is None or str(raw_periods).strip() == "":
            results.append(TeacherImportRowResult(
                row=row_num,
                teacher_name=teacher_name,
                status="invalid",
                reason="Max periods per day is empty.",
            ))
            continue

        try:
            max_periods = int(float(str(raw_periods)))
        except (ValueError, TypeError):
            results.append(TeacherImportRowResult(
                row=row_num,
                teacher_name=teacher_name,
                max_periods_per_day=None,
                status="invalid",
                reason=f"Invalid max periods value: '{raw_periods}'.",
            ))
            continue

        if max_periods < 1 or max_periods > 8:
            results.append(TeacherImportRowResult(
                row=row_num,
                teacher_name=teacher_name,
                max_periods_per_day=max_periods,
                status="invalid",
                reason=f"Max periods per day must be between 1 and 8 (got {max_periods}).",
            ))
            continue

        # ── Check intra-Excel duplicates ────────────────────
        name_key = teacher_name.lower()

        if name_key in seen_names:
            results.append(TeacherImportRowResult(
                row=row_num,
                teacher_name=teacher_name,
                max_periods_per_day=max_periods,
                status="skipped",
                reason=f"Duplicate of row {seen_names[name_key]} in this file.",
            ))
            continue

        # ── Check existing database records ─────────────────
        existing = _find_duplicate_teacher(db, teacher_name)

        if existing:
            results.append(TeacherImportRowResult(
                row=row_num,
                teacher_name=teacher_name,
                max_periods_per_day=max_periods,
                status="skipped",
                reason=f'A teacher named "{existing.teacher_name}" already exists.',
            ))
            seen_names[name_key] = row_num
            continue

        # ── Row is valid ────────────────────────────────────
        seen_names[name_key] = row_num
        valid_teachers.append({
            "teacher_name": teacher_name,
            "max_periods_per_day": max_periods,
        })
        results.append(TeacherImportRowResult(
            row=row_num,
            teacher_name=teacher_name,
            max_periods_per_day=max_periods,
            status="imported",
        ))

    wb.close()

    # ── Bulk insert ─────────────────────────────────────────
    imported_count = 0

    if valid_teachers:
        try:
            teacher_objects = [
                Teacher(**data) for data in valid_teachers
            ]
            db.add_all(teacher_objects)
            db.commit()
            imported_count = len(teacher_objects)
        except IntegrityError:
            db.rollback()
            # Fallback: insert one-by-one to salvage as many as possible
            for data in valid_teachers:
                try:
                    teacher = Teacher(**data)
                    db.add(teacher)
                    db.commit()
                    imported_count += 1
                except IntegrityError:
                    db.rollback()
                    # Mark this row as skipped in results
                    for r in results:
                        if (
                            r.teacher_name == data["teacher_name"]
                            and r.status == "imported"
                        ):
                            r.status = "skipped"
                            r.reason = "Database constraint violation (duplicate)."
                            break

    # ── Build response ──────────────────────────────────────
    total = len(results)
    skipped = sum(1 for r in results if r.status == "skipped")
    failed = sum(1 for r in results if r.status == "invalid")

    return TeacherImportResponse(
        total=total,
        imported=imported_count,
        skipped=skipped,
        failed=failed,
        rows=results,
    )


@router.get("/{teacher_id}", response_model=TeacherRead)
def get_teacher(
    teacher_id: int,
    db: Session = Depends(get_db),
) -> TeacherRead:

    teacher = (
        db.query(Teacher)
        .filter(Teacher.teacher_id == teacher_id)
        .first()
    )

    if not teacher:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Teacher not found",
        )

    return TeacherRead(
        teacher_id=teacher.teacher_id,
        teacher_name=teacher.teacher_name,
        max_periods_per_day=teacher.max_periods_per_day,
    )


@router.post(
    "",
    response_model=TeacherRead,
    status_code=status.HTTP_201_CREATED,
)
def create_teacher(
    payload: TeacherCreate,
    db: Session = Depends(get_db),
) -> TeacherRead:

    data = _dump_payload(payload)

    teacher_name = _normalise_teacher_name(
        data.get("teacher_name", "")
    )

    if not teacher_name:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Teacher name cannot be empty.",
        )

    duplicate = _find_duplicate_teacher(
        db,
        teacher_name,
    )

    if duplicate:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f'A teacher named "{duplicate.teacher_name}" already exists.',
        )

    data["teacher_name"] = teacher_name

    teacher = Teacher(**data)

    db.add(teacher)

    try:
        db.commit()
        db.refresh(teacher)

    except IntegrityError:
        db.rollback()

        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A teacher with this name already exists.",
        )

    return TeacherRead(
        teacher_id=teacher.teacher_id,
        teacher_name=teacher.teacher_name,
        max_periods_per_day=teacher.max_periods_per_day,
    )


@router.put(
    "/{teacher_id}",
    response_model=TeacherRead,
)
def update_teacher(
    teacher_id: int,
    payload: TeacherUpdate,
    db: Session = Depends(get_db),
) -> TeacherRead:

    teacher = (
        db.query(Teacher)
        .filter(Teacher.teacher_id == teacher_id)
        .first()
    )

    if not teacher:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Teacher not found",
        )

    data = _dump_payload(payload)

    if "teacher_name" in data and data["teacher_name"] is not None:

        teacher_name = _normalise_teacher_name(
            data["teacher_name"]
        )

        if not teacher_name:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Teacher name cannot be empty.",
            )

        duplicate = _find_duplicate_teacher(
            db,
            teacher_name,
            exclude_teacher_id=teacher_id,
        )

        if duplicate:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f'A teacher named "{duplicate.teacher_name}" already exists.',
            )

        data["teacher_name"] = teacher_name

    for field, value in data.items():
        setattr(teacher, field, value)

    db.add(teacher)

    try:
        db.commit()
        db.refresh(teacher)

    except IntegrityError:
        db.rollback()

        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A teacher with this name already exists.",
        )

    return TeacherRead(
        teacher_id=teacher.teacher_id,
        teacher_name=teacher.teacher_name,
        max_periods_per_day=teacher.max_periods_per_day,
    )


@router.delete(
    "/{teacher_id}",
    response_model=TeacherRead,
)
def delete_teacher(
    teacher_id: int,
    db: Session = Depends(get_db),
) -> TeacherRead:

    teacher = (
        db.query(Teacher)
        .filter(Teacher.teacher_id == teacher_id)
        .first()
    )

    if not teacher:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Teacher not found",
        )

    deleted = TeacherRead(
        teacher_id=teacher.teacher_id,
        teacher_name=teacher.teacher_name,
        max_periods_per_day=teacher.max_periods_per_day,
    )

    from app.models.models import (
        TeacherSubject,
        TeacherAvailability,
        Timetable,
    )

    db.query(TeacherSubject).filter(
        TeacherSubject.teacher_id == teacher_id
    ).delete(
        synchronize_session=False
    )

    db.query(TeacherAvailability).filter(
        TeacherAvailability.teacher_id == teacher_id
    ).delete(
        synchronize_session=False
    )

    db.query(Timetable).filter(
        Timetable.teacher_id == teacher_id
    ).delete(
        synchronize_session=False
    )

    db.delete(teacher)
    db.commit()

    return deleted