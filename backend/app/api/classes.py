import io
from typing import List

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from app.core.config import get_db
from app.models.models import Class as ClassModel
from app.schemas.class_ import ClassCreate, ClassRead, ClassUpdate, ClassImportRowResult, ClassImportResponse

router = APIRouter(prefix="/classes", tags=["classes"])


def _dump_payload(payload):
    """
    Support both Pydantic v1 (`dict`) and v2 (`model_dump`) for robustness.
    """
    if hasattr(payload, "model_dump"):
        return payload.model_dump(exclude_unset=True)
    return payload.dict(exclude_unset=True)


def _normalise_class_name(name: str) -> str:
    """Normalise whitespace around a class name."""
    return " ".join(name.strip().split())


def _find_duplicate_class(db: Session, class_name: str):
    """Find a class with the same name, ignoring case and whitespace."""
    normalised = _normalise_class_name(class_name)
    return (
        db.query(ClassModel)
        .filter(func.lower(func.trim(ClassModel.class_name)) == normalised.lower())
        .first()
    )


@router.get("", response_model=List[ClassRead])
def get_classes(db: Session = Depends(get_db)) -> List[ClassRead]:
    classes = db.query(ClassModel).all()
    return [ClassRead(class_id=c.class_id, class_name=c.class_name) for c in classes]


@router.get("/import/template")
def download_class_import_template():
    """Generate and return an Excel template for class imports."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Classes"

    ws.append(["class_name"])
    ws.append(["CS-A"])
    ws.append(["AI-B"])

    ws.column_dimensions["A"].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=class_import_template.xlsx"
        },
    )


@router.post("/import", response_model=ClassImportResponse)
def import_classes(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> ClassImportResponse:
    """
    Bulk-import classes from an Excel (.xlsx) file.

    Valid rows are inserted. Invalid and duplicate rows are
    skipped and reported — they do NOT block other rows.
    """

    # ── Validate file type ──────────────────────────────────
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are accepted.",
        )

    # ── Read workbook ───────────────────────────────────────
    try:
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read the Excel file. Please upload a valid .xlsx file.",
        )

    ws = wb.active
    if ws is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The workbook has no active sheet.",
        )

    rows_iter = ws.iter_rows(values_only=True)

    # ── Validate headers ────────────────────────────────────
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The Excel file is empty.",
        )

    headers = [
        str(h).strip().lower() if h is not None else ""
        for h in header_row
    ]

    if "class_name" not in headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required column: class_name. Found: {', '.join(headers) or '(none)'}.",
        )

    name_idx = headers.index("class_name")

    # ── Read and validate data rows ─────────────────────────
    results: list[ClassImportRowResult] = []
    valid_classes: list[dict] = []
    seen_names: dict[str, int] = {}  # normalised_lower -> first row number

    for row_num_offset, row in enumerate(rows_iter):
        row_num = row_num_offset + 2  # 1-indexed, header is row 1

        raw_name = row[name_idx] if name_idx < len(row) else None

        # ── Validate class_name ─────────────────────────────
        if raw_name is None or str(raw_name).strip() == "":
            results.append(ClassImportRowResult(
                row=row_num,
                class_name=str(raw_name or ""),
                status="invalid",
                reason="Class name is empty.",
            ))
            continue

        class_name = _normalise_class_name(str(raw_name))

        if not class_name:
            results.append(ClassImportRowResult(
                row=row_num,
                class_name=str(raw_name),
                status="invalid",
                reason="Class name is empty after trimming.",
            ))
            continue

        if len(class_name) > 100:
            results.append(ClassImportRowResult(
                row=row_num,
                class_name=class_name[:50] + "...",
                status="invalid",
                reason="Class name exceeds 100 characters.",
            ))
            continue

        # ── Check intra-Excel duplicates ────────────────────
        name_key = class_name.lower()

        if name_key in seen_names:
            results.append(ClassImportRowResult(
                row=row_num,
                class_name=class_name,
                status="skipped",
                reason=f"Duplicate of row {seen_names[name_key]} in this file.",
            ))
            continue

        # ── Check existing database records ─────────────────
        existing = _find_duplicate_class(db, class_name)

        if existing:
            results.append(ClassImportRowResult(
                row=row_num,
                class_name=class_name,
                status="skipped",
                reason=f'A class named "{existing.class_name}" already exists.',
            ))
            seen_names[name_key] = row_num
            continue

        # ── Row is valid ────────────────────────────────────
        seen_names[name_key] = row_num
        valid_classes.append({"class_name": class_name})
        results.append(ClassImportRowResult(
            row=row_num,
            class_name=class_name,
            status="imported",
        ))

    wb.close()

    # ── Bulk insert ─────────────────────────────────────────
    imported_count = 0

    if valid_classes:
        try:
            class_objects = [ClassModel(**data) for data in valid_classes]
            db.add_all(class_objects)
            db.commit()
            imported_count = len(class_objects)
        except IntegrityError:
            db.rollback()
            # Fallback: insert one-by-one to salvage as many as possible
            for data in valid_classes:
                try:
                    obj = ClassModel(**data)
                    db.add(obj)
                    db.commit()
                    imported_count += 1
                except IntegrityError:
                    db.rollback()
                    for r in results:
                        if (
                            r.class_name == data["class_name"]
                            and r.status == "imported"
                        ):
                            r.status = "skipped"
                            r.reason = "Database constraint violation (duplicate)."
                            break

    # ── Build response ──────────────────────────────────────
    total = len(results)
    skipped = sum(1 for r in results if r.status == "skipped")
    failed = sum(1 for r in results if r.status == "invalid")

    return ClassImportResponse(
        total=total,
        imported=imported_count,
        skipped=skipped,
        failed=failed,
        rows=results,
    )


@router.get("/{class_id}", response_model=ClassRead)
def get_class(class_id: int, db: Session = Depends(get_db)) -> ClassRead:
    class_obj = db.query(ClassModel).filter(ClassModel.class_id == class_id).first()
    if not class_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Class not found",
        )
    return ClassRead(class_id=class_obj.class_id, class_name=class_obj.class_name)


@router.post("", response_model=ClassRead, status_code=status.HTTP_201_CREATED)
def create_class(payload: ClassCreate, db: Session = Depends(get_db)) -> ClassRead:
    data = _dump_payload(payload)
    class_obj = ClassModel(**data)
    db.add(class_obj)
    db.commit()
    db.refresh(class_obj)
    return ClassRead(class_id=class_obj.class_id, class_name=class_obj.class_name)


@router.put("/{class_id}", response_model=ClassRead)
def update_class(
    class_id: int, payload: ClassUpdate, db: Session = Depends(get_db)
) -> ClassRead:
    class_obj = db.query(ClassModel).filter(ClassModel.class_id == class_id).first()
    if not class_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Class not found",
        )

    data = _dump_payload(payload)
    for field, value in data.items():
        setattr(class_obj, field, value)

    db.add(class_obj)
    db.commit()
    db.refresh(class_obj)
    return ClassRead(class_id=class_obj.class_id, class_name=class_obj.class_name)


@router.delete("/{class_id}", response_model=ClassRead)
def delete_class(class_id: int, db: Session = Depends(get_db)) -> ClassRead:
    class_obj = db.query(ClassModel).filter(ClassModel.class_id == class_id).first()
    if not class_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Class not found",
        )

    deleted = ClassRead(class_id=class_obj.class_id, class_name=class_obj.class_name)

    from app.models.models import Subject, TeacherSubject, Timetable
    subject_ids = [s.subject_id for s in db.query(Subject).filter(Subject.class_id == class_id).all()]
    if subject_ids:
        db.query(TeacherSubject).filter(TeacherSubject.subject_id.in_(subject_ids)).delete(synchronize_session=False)
        db.query(Timetable).filter(Timetable.subject_id.in_(subject_ids)).delete(synchronize_session=False)
        db.query(Subject).filter(Subject.class_id == class_id).delete(synchronize_session=False)
    db.query(Timetable).filter(Timetable.class_id == class_id).delete(synchronize_session=False)

    db.delete(class_obj)
    db.commit()
    return deleted

